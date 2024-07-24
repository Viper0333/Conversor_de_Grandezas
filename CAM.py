from openpyxl import load_workbook
import os
import datetime
import pandas as pd

# Diretório do arquivo .py atual
diretorio_conversor = os.path.dirname(os.path.abspath(__file__)) 

# Data e hora atuais
data_e_hora_atuais = datetime.datetime.now() 
data_formatada = data_e_hora_atuais.strftime("%d_%m_%Y") 

# Cria a pasta para arquivos convertidos com a data atual
pasta_ABS = os.path.join(diretorio_conversor, 'dados')
arquivos_convertidos = os.path.join(diretorio_conversor, 'Arquivos_convertidos')
nome_pasta = f"{data_formatada}_Arquivos_Milimetro" 
caminho_pasta = os.path.join(arquivos_convertidos, nome_pasta)

if not os.path.exists(arquivos_convertidos):
    os.makedirs(arquivos_convertidos)

if not os.path.exists(caminho_pasta):
    os.makedirs(caminho_pasta)

# Loop para arquivos na pasta
for arquivo_amper in os.listdir(pasta_ABS):
    caminho_arquivo = os.path.join(pasta_ABS, arquivo_amper)
    nome_arquivo_xlsx = os.path.splitext(arquivo_amper)[0] + ".xlsx"  
    caminho_arquivo_xlsx = os.path.join(caminho_pasta, nome_arquivo_xlsx)
    
    # Verifica se o arquivo é um arquivo com a extensão CSV
    if arquivo_amper.endswith(".csv"):
        dados = pd.read_csv(caminho_arquivo, sep=";") 
        print(arquivos_convertidos)
        # Salva o DataFrame como um arquivo XLSX na pasta nome_pasta
        dados.to_excel(caminho_arquivo_xlsx, index=False, engine='xlsxwriter')
        
    else:
        print(f"O arquivo {arquivo_amper} não é um arquivo CSV. Pulando para o próximo arquivo.")
        continue

    # Carrega o arquivo Excel
    wb = load_workbook(caminho_arquivo_xlsx)  
    pagina1 = wb.active

    # Adiciona cabeçalhos nas colunas F, G e H
    pagina1['F1'] = "Nivel_em_Amperes"
    pagina1['G1'] = "Nivel_em_Milimetros"
    pagina1['H1'] = "Nivel_em_Centimetros"

    bool_valido = [0]
    contador = 0

    # Itera sobre cada linha na coluna F
    for row in pagina1.iter_rows(min_col=6, max_col=6, values_only=False):
        for cell in row:
            cell_value = cell.value  # Valor da célula
            cell_row = cell.row # Número da linha da célula
            quantidade_linhas = pagina1.max_row - 1 # Obter a quantidade de linhas na planilha

            if cell_value == 0:
                contador += 1
            
            if str(cell_value).isdigit() and int(cell_value) != 0:
                # Calcula os valores para as colunas G e H
                resultado = int(cell_value) * 1450 - 340000
                milimetros = resultado / 1600

                # Escreve os valores nas colunas G e H
                pagina1.cell(row=cell_row, column=7, value=round(milimetros, 1))
                pagina1.cell(row=cell_row, column=8, value=round(milimetros / 10, 1))

                

                bool_valido = 1
                contador += 1 
                                
                #print("Boleano é:",bool_valido)
                #print("Contador é:",contador)
                #print("Quantidade de linhas é:",quantidade_linhas)

            if contador == quantidade_linhas and bool_valido == 1:
                #print("Contador igual a quantidade de linhas",contador,"=",quantidade_linhas)
                os.remove(caminho_arquivo)
                
        
        # Salva o arquivo Excel com as alterações
        wb.save(caminho_arquivo_xlsx)